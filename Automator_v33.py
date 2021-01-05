import pandas as pd
import numpy as np
import glob
import datetime
import openpyxl
import sys
from openpyxl import Workbook
from datetime import datetime, timedelta

import time
import string
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from copy import copy
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import load_workbook

import warnings
warnings.filterwarnings(action='ignore')

def clr_to_class_v3(c):
    clr_to_class={}
    clr_to_class['4'] = '정기외박'
    clr_to_class['6'] = '교육'
    clr_to_class['9'] = '특별외박'
    clr_to_class['FFFF9999'] = '정기휴가'
    clr_to_class['FFFFFF00'] = '병가'
    clr_to_class['1'] = '휴직'
    clr_to_class['FF7030A0'] = '공가'
    try :
        return clr_to_class[c]
    except :
        print(c)
        raise ValueError("기존 영외활동 컬러와 일치하지 않습니다. 확인하시기 바랍니다.")

def date_column_cell(df):
  alphabets = string.ascii_uppercase
  date_list = list(df.index)
  column_list = [i for i in range(len(df.columns))]
  ref_table = pd.DataFrame(index = date_list, columns = column_list)
  row_integer = [i for i in range(len(date_list))]
  for i in column_list :
    ref_table.loc[:,i] = [alphabets[i+2] + str(num_row+2) for num_row in row_integer]
  return ref_table

def SpecialOutExtract(df,start_date,end_date):
  contents = df[(df.index >= start_date) & (df.index <= end_date)].specialty.values
  rst = ''
  for i in range(len(contents)):
    if contents[i] != '':
      rst += contents[i]
      if i != len(contents)-1:
        rst += '&'
  return rst

def SpecialOutExtract_v2(df,start_date,end_date):
  test_df = df[(df.index >= start_date) & (df.index <= end_date)]
  if len(test_df.index) != len(set(list(test_df.index))):
    if len(df[(df.index==end_date)]) == 2:
      contents = df[(df.index >= start_date) & (df.index <= end_date)].specialty.values[:-1]
      rst = ''
      for i in range(len(contents)):
        if contents[i] != '':
          rst += contents[i]
          if i != len(contents)-1:
            rst += '&'
    else :
      if len(df[(df.index==start_date)])==2:
        contents = df[(df.index >= start_date) & (df.index <= end_date)].specialty.values[1:]
        rst = ''
        for i in range(len(contents)):
          if contents[i] != '':
            rst += contents[i]
            if i != len(contents)-1:
              rst += '&'
  else :
    contents = df[(df.index >= start_date) & (df.index <= end_date)].specialty.values
    rst = ''
    for i in range(len(contents)):
      if contents[i] != '':
        rst += contents[i]
        if i != len(contents)-1:
          rst += '&'
  return rst


def DoubleChecker(df):
  for i in range(len(df)):
    if len(df.iloc[i,0].split('/'))>1:
      return True
  return False

def DoubleSpliter_v3(df):
  df.columns = ['name']
  df_idx = [str(idx)[:10] for idx in df.index]
  df.index = df_idx
  sub_df = df[df['name'].str.contains('/')]
  if len(sub_df) > 0 :
    idx_db_name_list = []
    idx_ls = [str(idx)[:10] for idx in sub_df.index]
    name_ls = sub_df.name.values
    for i in range(len(sub_df)):
      idx_db_name_list.append([idx_ls[i],name_ls[i].split('/')[-1]]) # to be added
    df['name'] = df.loc[:,'name'].map(lambda x : x.split('/')[0] if len(x.split('/'))>1 else x)   

    temp_df = pd.DataFrame()
    temp_ls = []
    for i in range(len(idx_db_name_list)):
      if i != len(idx_db_name_list)-1:
        dt = idx_db_name_list[i][0]
        next_dt = idx_db_name_list[i+1][0]
        nm = idx_db_name_list[i][1]
        next_nm = idx_db_name_list[i+1][1]
        temp_df = pd.concat([temp_df,pd.DataFrame(data=[nm],columns=['name'],index=[dt])])
        if (nm != next_nm) or (datetime(int(next_dt[:4]),int(next_dt[5:7]),int(next_dt[8:10])) - datetime(int(dt[:4]),int(dt[5:7]),int(dt[8:10])) != timedelta(1)):
          temp_ls.append(temp_df)
          temp_df = pd.DataFrame()
      else :
        dt = idx_db_name_list[i][0]
        nm = idx_db_name_list[i][1]
        temp_df = pd.concat([temp_df,pd.DataFrame(data=[nm],columns=['name'],index=[dt])])
        temp_ls.append(temp_df)
        temp_df = pd.DataFrame()
              
    for i in range(len(temp_ls)):
      duplicated_df = temp_ls[i]
      start_dt = duplicated_df.index[0]
      end_dt = duplicated_df.index[-1]
      first_df = df[lambda x : x.index <= end_dt]
      last_df = df[lambda x : x.index > end_dt]
      df = pd.concat([first_df,duplicated_df])
      df = pd.concat([df,last_df]) 
    
  else :
    return 
  df_idx = [str(idx)[:10] for idx in df.index]
  df.index = df_idx
  return df

def logic_v6(df,ith,sheet,ref):
  df_idx = [str(idx)[:10] for idx in df.index]
  df.index = df_idx
  

  if len(df) == 0 :
    return
  
  if DoubleChecker(df):
    df = DoubleSpliter_v3(df)
  else :
    pass
  
  # 1 Column Unit Operation
  color_ls = []
  for d in df.index:
    clr = str(sheet[ref.loc[str(d)[:10],ith]].fill.start_color.index)
    if clr == '8': #
      clr = '4'
    if clr == 'FFFF5D5D': #휴가 /로 중첩해서 쓴 색깔을 뜻하는것.
      clr = 'FFFF9999'
    color_ls.append(clr)
  
  df['color'] = color_ls
  df['specialty'] = df.iloc[:,0].map(lambda x : '(' + x.split('(')[1] if len(x.split('('))>1 else '')
  df.iloc[:,0] = pd.DataFrame(df.iloc[:,0].map(lambda x : x.split('(')[0]))
  ref_ls = []
  for i in range(len(df)):
    ref_ls.append([df.index[i],df.iloc[i,0],df.iloc[i,1]])
    
  dates_ls = list(df.index)
  total_ls = []
  sub_ls = []
  is_first = True
  for idx, dt in enumerate(dates_ls) :
    try :
      if idx != len(dates_ls)-1 :
        next_dt = dates_ls[idx+1]
        if dt != next_dt :
          if ref_ls[idx][1] == ref_ls[idx+1][1] :
            if ref_ls[idx][2] == ref_ls[idx+1][2] : 
              if datetime(int(next_dt[:4]),int(next_dt[5:7]),int(next_dt[8:10])) - datetime(int(dt[:4]),int(dt[5:7]),int(dt[8:10])) == timedelta(1):
                if is_first :
                  sub_ls.append(ref_ls[idx][1])
                  sub_ls.append(ref_ls[idx][2])
                  sub_ls.append(dt)
                  is_first = False
                else :
                  pass
              else : 
                sub_ls.append(dt)
                sub_ls.append(SpecialOutExtract_v2(df,sub_ls[-2],sub_ls[-1]))
                total_ls.append(sub_ls)
                sub_ls = []
                is_first = True
            else :
              sub_ls.append(dt)
              sub_ls.append(SpecialOutExtract_v2(df,sub_ls[-2],sub_ls[-1]))
              total_ls.append(sub_ls)
              sub_ls = []
              is_first = True
          else :
            sub_ls.append(dt)
            sub_ls.append(SpecialOutExtract_v2(df,sub_ls[-2],sub_ls[-1]))
            total_ls.append(sub_ls)
            sub_ls = []
            is_first = True
        else :
          sub_ls.append(dt)
          sub_ls.append(SpecialOutExtract_v2(df,sub_ls[-2],sub_ls[-1]))
          total_ls.append(sub_ls)
          sub_ls = []
          is_first = True
      else :
        sub_ls.append(dt)
        sub_ls.append(SpecialOutExtract_v2(df,sub_ls[-2],sub_ls[-1]))
        total_ls.append(sub_ls)
        sub_ls = []
        is_first = True
    except :
      sub_ls = []
      pass
  return total_ls

def clr_to_class(c):
  clr_to_class = {}
  clr_to_class['4'] = '정기외박'
  clr_to_class['9'] = '특별외박'
  clr_to_class['FFFF9999'] = '정기휴가'
  try :
    return clr_to_class[c]
  except:
    return '병가'

def clr_to_class_v2(c):
  clr_to_class={}
  clr_to_class['4'] = '정기외박'
  clr_to_class['6'] = '교육'
  clr_to_class['9'] = '특별외박'
  clr_to_class['FFFF9999'] = '정기휴가'
  clr_to_class['FFFFFF00'] = '병가'
  try :
    return clr_to_class[c]
  except :
    print(c)
    raise ValueError("기타/공가를 원하는거면 버전을 업데이트 하세요.")


def ChungwonConverter(df):
  temp_idx = list(df[df['Specialty'].str.contains('청원')].index)
  df.loc[temp_idx,'Class'] = '청원휴가'
  return df

def NameToTeamConverter(nm,nmteamdict):
  return nmteamdict[nm]

def PeriodExtract(df):
  period_ls = []
  for i in range(len(df)):
    stdt = df.loc[i,'StartDate']
    eddt = df.loc[i,'EndDate']
    prd = stdt[5:7] + '.' + stdt[8:10] + ' ~ ' + eddt[5:7] + '.' + eddt[8:10]
    period_ls.append(prd)
  df['기간'] = period_ls
  return df

def Stacking_v3(df,db,sheet,ref):
  total_table = pd.DataFrame(columns=['Name','Class','StartDate','EndDate','Specialty'])
  for i, cl in enumerate(df.columns):
    temp_output = pd.DataFrame(data=logic_v6(df[[cl]].dropna(axis=0),i,sheet,ref),columns=['Name','Class','StartDate','EndDate','Specialty'])
    total_table = pd.concat([total_table,temp_output])
    #print(temp_output)
  total_table['Class'] = total_table['Class'].map(lambda x : clr_to_class_v3(x)) #v2or없애기
  total_table.reset_index(inplace=True,drop=True)
  nmteamdict = {}
  for i in range(len(db)):
    nmteamdict[db.loc[i,'성명']] = str(db.loc[i,'소대'])
  total_table['Team'] = total_table['Name'].map(lambda x : NameToTeamConverter(x,nmteamdict))
  total_table = PeriodExtract(total_table)
  total_table.reset_index(inplace=True, drop=True)
  total_table = ChungwonConverter(total_table)
  return total_table

def SortByTeam(df):
  #print('Sorting..')
  df = df.sort_values(by=['소 대'],ascending=True)
  return df#.sort_values(by=['소 대'],inplace=True)

def Sagoja_v5(table, dt, db,IsCompress=True):
  sub_table = table[(table['StartDate']<=dt) & (table['EndDate']>=dt)]
  rst_df = pd.DataFrame(columns=['연번','구분','소대','소대','성명','기간','종별','비고'])
  sub_table.reset_index(inplace=True,drop=True)
  sub_table['Team'] = sub_table['Team'].map(lambda x : x + '소대' if len(x)==1 else '본부소대')
  
  type_ls = []
  for i in range(len(sub_table)):
    #구분
    tpcls = sub_table.loc[i,'Class']
    if '휴가' in tpcls :
      type_ls.append('휴 가')
    elif '정기외박' in tpcls :
      type_ls.append('외 박')
    elif '특별외박' in tpcls :
      type_ls.append('특 박')
    elif '병가' in tpcls :
      type_ls.append('병 가')
    elif '교육' in tpcls :
      type_ls.append('교 육')
    else :
      type_ls.append('기 타')
      print(" Problem : ", tpcls)
      raise ValueError('Check this out !!!')
  sub_table['구 분'] = type_ls
  sub_table.rename(columns={'Name':'성 명','Team':'소 대','Class':'종 별','Specialty':'비 고','기간':'기 간'},inplace=True)
  sub_table.drop(['StartDate','EndDate'],axis=1,inplace=True)
  sub_table = sub_table[['구 분','소 대','성 명','기 간','종 별','비 고']]
  if IsCompress == True :
    pass
  else :
    full_sagoja = sub_table.copy()
  rst_table = pd.DataFrame(columns = ['구 분','소 대','성 명','기 간','종 별','비 고'])
  heuga_df = SortByTeam(sub_table[lambda x : x['구 분']=='휴 가'])
  try:
    heuga_df = Compress_v1(heuga_df)
    heuga_df.iloc[1:,0] = ''
  except:
    pass
  
  whybak_df = SortByTeam(sub_table[lambda x : x['구 분']=='외 박'])
  try:
    whybak_df = Compress_v1(whybak_df)
    whybak_df.iloc[1:,0] = ''
  except:
    pass
  specialbak_df = SortByTeam(sub_table[lambda x : x['구 분']=='특 박'])
  try:
    specialbak_df = Compress_v1(specialbak_df)
    specialbak_df.iloc[1:,0] = ''
  except:
    pass
  sickbak_df = SortByTeam(sub_table[lambda x : x['구 분']=='병 가'])
  try:
    sickbak_df = Compress_v1(sickbak_df)
    sickbak_df.iloc[1:,0] = ''
  except:
    pass
  edubak_df = SortByTeam(sub_table[lambda x : x['구 분']=='교 육'])
  try :
    edubak_df = Compress_v1(edubak_df)
    edubak_df.iloc[1:,0] = ''
  except :
    pass
  guitar_df = SortByTeam(sub_table[lambda x : x['구 분']=='기 타'])
  try:
    guitar_df = Compress_v1(guitar_df)
    guitar_df.iloc[1:,0] = ''
  except:
    pass
  rst_table = pd.concat([rst_table,heuga_df])
  rst_table = pd.concat([rst_table,whybak_df])
  rst_table = pd.concat([rst_table,specialbak_df])
  rst_table = pd.concat([rst_table,sickbak_df])
  rst_table = pd.concat([rst_table,edubak_df])
  rst_table = pd.concat([rst_table,guitar_df])
  rst_table.reset_index(drop=True,inplace=True)
  yeonbeon_ls = [i+1 for i in rst_table.index]
  rst_table['연 번'] = yeonbeon_ls
  if IsCompress == True :
    return rst_table[['연 번','구 분','소 대','성 명','기 간','종 별','비 고']]
  else :
    full_sagoja['연 번'] = yeonbeon_ls
    return full_sagoja[['연 번','구 분','소 대','성 명','기 간','종 별','비 고']]

def Compress_v1(df,col_nm='소 대',val_ls=['1소대','2소대','3소대','본부소대']):
  rst_table = pd.DataFrame(columns = ['구 분','소 대','성 명','기 간','종 별','비 고'])
  heuga_df = df[lambda x : x[col_nm]==val_ls[0]]
  try:
    heuga_df.iloc[1:,1] = ''
  except:
    pass
  whybak_df = df[lambda x : x[col_nm]==val_ls[1]]
  try:
    whybak_df.iloc[1:,1] = ''
  except:
    pass
  specialbak_df = df[lambda x : x[col_nm]==val_ls[2]]
  try:
    specialbak_df.iloc[1:,1] = ''
  except:
    pass
  sickbak_df = df[lambda x : x[col_nm]==val_ls[3]]
  try:
    sickbak_df.iloc[1:,1] = ''
  except:
    pass
  rst_table = pd.concat([rst_table,heuga_df])
  rst_table = pd.concat([rst_table,whybak_df])
  rst_table = pd.concat([rst_table,specialbak_df])
  rst_table = pd.concat([rst_table,sickbak_df])
  rst_table.reset_index(drop=True,inplace=True)
  return rst_table[['구 분','소 대','성 명','기 간','종 별','비 고']]

def SagojaDesign_v3(df,dt,wb):
  ws = wb['사고자신임']
  a1 = ws['A1']
  a1.value = '□ 주 요 사 고 내 용'
  a1.font = Font(name='맑은 고딕',size=27,bold=True) 
  g2 = ws['G2']
  g2.value = '일자 : '+ dt
  sub1 = Font(name='맑은 고딕',size=11)
  g2.font = sub1
  alphabets = string.ascii_uppercase

  for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

  alnmt = Alignment(horizontal="center",vertical="center")
  all_brdr = Border(left=Side(style="thin", color="000000"),
                  right=Side(style='thin', color="000000"),
                  top=Side(style="thin", color="000000"),
                  bottom=Side(style="thin", color="000000"),
                  diagonal=Side(style="thin", color="000000"),
                  diagonal_direction=0,
                  outline=Side(style="thin", color="000000"),
                  vertical=Side(style="thin", color="000000"),
                  horizontal=Side(style="thin", color="000000"))
  
  allbutup_brdr = Border(left=Side(style="thin", color="000000"),
                  right=Side(style='thin', color="000000"),
                  bottom=Side(style="thin", color="000000"),
                  diagonal=Side(style="thin", color="000000"),
                  diagonal_direction=0,                  outline=Side(style="thin", color="000000"),
                  vertical=Side(style="thin", color="000000"),
                  horizontal=Side(style="thin", color="000000"))
  
  side_brdr = Border(left=Side(style="thin", color="000000"),
                  right=Side(style='thin', color="000000"),
                  diagonal=Side(style="thin", color="000000"),
                  diagonal_direction=0,
                  outline=Side(style="thin", color="000000"),
                  vertical=Side(style="thin", color="000000"),
                  horizontal=Side(style="thin", color="000000"))

  allbutbtm_brdr = Border(left=Side(style="thin", color="000000"),
                  right=Side(style='thin', color="000000"),
                  top=Side(style="thin", color="000000"),
                  diagonal=Side(style="thin", color="000000"),
                  diagonal_direction=0,
                  outline=Side(style="thin", color="000000"),
                  vertical=Side(style="thin", color="000000"),
                  horizontal=Side(style="thin", color="000000"))

  for ap in range(len(alphabets[:7])):
    for nb in range(3,3+len(df)+1):
      ws[alphabets[:7][ap]+str(nb)].alignment = alnmt
      ws[alphabets[:7][ap]+str(nb)].font = Font(name='맑은 고딕',size=11,bold=True)
      if (nb-3 >= 1) & (nb-3 <= len(df)-1) & (alphabets[:7][ap] in ['B']):#,'C']) :
        if nb-3 != len(df)-1:
          if df.loc[nb-3,'구 분'] == '':
            ws[alphabets[:7][ap]+str(nb+1)].border = side_brdr
          else :
            ws[alphabets[:7][ap]+str(nb+1)].border = allbutbtm_brdr
        else :
          if df.loc[nb-3,'구 분'] == '':
            ws[alphabets[:7][ap]+str(nb+1)].border = allbutup_brdr
          else :
            ws[alphabets[:7][ap]+str(nb+1)].border = all_brdr  
      elif (nb-3 >= 1) & (nb-3 <= len(df)-1) & (alphabets[:7][ap] in ['C']):
        if nb-3 != len(df)-1:
          if df.loc[nb-3,'소 대'] == '':
            ws[alphabets[:7][ap]+str(nb+1)].border = side_brdr
          else :
            ws[alphabets[:7][ap]+str(nb+1)].border = allbutbtm_brdr
        else :
          if df.loc[nb-3,'소 대'] == '':
            ws[alphabets[:7][ap]+str(nb+1)].border = allbutup_brdr
          else :
            ws[alphabets[:7][ap]+str(nb+1)].border = all_brdr
      else :
        if (alphabets[:7][ap] not in ['B','C']):
          ws[alphabets[:7][ap]+str(nb)].border = all_brdr
        else :
          if nb == 3:
            ws[alphabets[:7][ap]+str(nb)].border = all_brdr
  ws['B4'].border = allbutbtm_brdr
  ws.column_dimensions['A'].width = 6.7#6.0
  ws.column_dimensions['B'].width = 8.7#8.0
  ws.column_dimensions['C'].width = 8.7#8.0
  ws.column_dimensions['D'].width = 11.1#10.40
  ws.column_dimensions['E'].width = 17.6#16.90
  ws.column_dimensions['F'].width = 11.1#10.40
  ws.column_dimensions['G'].width = 29.0#22.1#21.40

  for i in range(1,3+len(df)+1):
    if  i==1 :
      ws.row_dimensions[i].height = 40.20 #39.00
    elif i==2 :
      ws.row_dimensions[i].height = 17.40 #16.50
    else :
      ws.row_dimensions[i].height = 24.00
  ws.print_area = 'A1:G'+str(3+len(df))
  return 'Sagoja designing is finished~'

def checklist_v5(sagoja_df,dt,wb):
  ws = wb['영외활동점검부']
  alphabets = string.ascii_uppercase
  ws['N1'] = '일자 : '+ dt 
  another_dt = '~ '+dt[5:7]+'.'+dt[8:10]
  toberemoved_ls = list(sagoja_df[sagoja_df['기 간'].str.contains(another_dt)].index)
  sagoja_df = sagoja_df[~sagoja_df.index.isin(toberemoved_ls)]
  sagoja_df.reset_index(drop=True, inplace=True)
  yeonbeon_ls = [i+1 for i in range(len(sagoja_df))]
  sagoja_df['연 번'] = yeonbeon_ls
  sagoja_df['소 대'] = sagoja_df.loc[:,'소 대'].map(lambda x : x.split('소대')[0])
  
  jungiheuga = len(sagoja_df[lambda x : x['종 별']=='정기휴가'])
  jungiwhybak = len(sagoja_df[lambda x : x['종 별']=='정기외박'])
  specialbak = len(sagoja_df[lambda x : x['종 별']=='특별외박'])
  chungwonheuga = len(sagoja_df[lambda x : x['종 별']=='청원휴가'])
  sickbak = len(sagoja_df[lambda x : x['종 별']=='병가'])
  edubak = len(sagoja_df[lambda x : x['종 별']=='교육'])
  guitar = len(sagoja_df[lambda x : x['종 별']=='기타'])
  ws['E4'].value = '정기휴가 '+str(jungiheuga)+'명, '+'청원휴가 '+str(chungwonheuga)+'명, '+'정기외박 '+str(jungiwhybak)+'명, '+'특별외박 '+str(specialbak)+'명, '+'병가 '+str(sickbak)+'명, '+'기타 '+str(guitar)+'명'

  all_brdr = Border(left=Side(style="thin", color="000000"),
                    right=Side(style='thin', color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"),
                    diagonal=Side(style="thin", color="000000"),
                    diagonal_direction=0,
                    outline=Side(style="thin", color="000000"),
                    vertical=Side(style="thin", color="000000"),
                    horizontal=Side(style="thin", color="000000"))
  alnmt = Alignment(horizontal="center",vertical="center")
  for i in range(len(sagoja_df)):
    ws['B'+str(i+11)] = sagoja_df.loc[i,'연 번']
    ws['B'+str(i+11)].alignment = alnmt
    ws['C'+str(i+11)] = sagoja_df.loc[i,'소 대']
    ws['C'+str(i+11)].alignment = alnmt
    ws['D'+str(i+11)] = sagoja_df.loc[i,'성 명']
    ws['D'+str(i+11)].alignment = alnmt
    ws['E'+str(i+11)] = sagoja_df.loc[i,'기 간']
    ws['E'+str(i+11)].alignment = alnmt
    ws['F'+str(i+11)] = sagoja_df.loc[i,'종 별']
    ws['F'+str(i+11)].alignment = alnmt
  for ap in range(len(alphabets[1:15])):
    for i in range(len(sagoja_df)):
      ws[alphabets[1:15][ap]+str(i+11)].border = all_brdr
  thin = Side(border_style="thin",color="000000")
  thick = Side(border_style="thick", color="000000")
  
  for rangee in ws.merged_cells.ranges:
    if (len(str(rangee).split(' ')[-1])>5) or ('B6' in str(rangee).split(' ')[-1]) or ('B7' in str(rangee).split(' ')[-1]) or ('G9' in str(rangee).split(' ')[-1]) or ('I9' in str(rangee).split(' ')[-1]) or ('L9' in str(rangee).split(' ')[-1]):
      style_range(ws, str(rangee), border=all_brdr)#,alignment=alnmt)
  for cl in range(len(alphabets[1:15])):
    ws[alphabets[cl+1]+str(2)].border = Border(top=thick,bottom=thick)
  ws.print_area = 'A1:O'+str(11+len(sagoja_df)-1)
  return "Checklist is done"

def style_range(ws, cell_range, border=Border(), fill=None, font=None, 
alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

def createNewWorkbook(manyWb):
    for wb in manyWb:
        for sheetName in wb.sheetnames:
            o = theOne.create_sheet(sheetName)
            safeTitle = o.title
            copySheet(wb[sheetName],theOne[safeTitle])

def copySheet(sourceSheet,newSheet):
    for row in sourceSheet.rows:
        for cell in row:
            newCell = newSheet.cell(row=cell.row, column=cell.col_idx,
                    value= cell.value)
            if cell.has_style:
                newCell.font = copy(cell.font)
                newCell.border = copy(cell.border)
                newCell.fill = copy(cell.fill)
                newCell.number_format = copy(cell.number_format)
                newCell.protection = copy(cell.protection)
                newCell.alignment = copy(cell.alignment)
    for col in sourceSheet.column_dimensions:
      newSheet.column_dimensions[col] = sourceSheet.column_dimensions[col]
    for row in sourceSheet.row_dimensions:
      newSheet.row_dimensions[row] = sourceSheet.row_dimensions[row]
    newSheet.merged_cells = copy(sourceSheet.merged_cells)

"""filesInput = sys.argv[1:]
theOneFile = filesInput.pop(-1)
myfriends = [ load_workbook(f) for f in filesInput ]

#try this if you are bored
#myfriends = [ openpyxl.load_workbook(f) for k in range(200) for f in filesInput ]

theOne = Workbook()
del theOne['Sheet'] #We want our new book to be empty. Thanks.
createNewWorkbook(myfriends)
theOne.save(theOneFile)"""

def PatrolChart_v5(dt,wb):
  t = ['월','화','수','목','금','토','일']
  r = datetime(int(dt[:4]),int(dt[5:7]),int(dt[8:10])).weekday()
  input_dt = dt[:4]+'년 '+dt[5:7]+'월 '+dt[8:10]+'일'+'\n'+t[r]+'요일'
  ws = wb['순찰표']
  col_dims = ws.column_dimensions
  row_dims = ws.row_dimensions

  ws['B5'].value = '( 지 하 ) 순찰표\n'+input_dt
  ws['H5'].value = '( 무 기 고 ) 순찰표\n'+input_dt
  ws['B31'].value = '( 서 측 계 단 ) 순찰표\n'+input_dt
  ws['B56'].value = '(3소대 내무실) 순찰표\n'+input_dt
  ws['H56'].value = '(2소대 내무실) 순찰표\n'+input_dt
  ws['B87'].value = '(1소대 내무실) 순찰표\n'+input_dt
  ws['H87'].value = '(본부소대 내무실) 순찰표\n'+input_dt

  thick = Side(border_style="thick", color="000000")
  thin = Side(border_style="thin",color="000000")
  border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)
  border_thick = Border(top=thick, left=thick, right=thick, bottom=thick)
  double = Side(border_style="double", color="000000")

  for rangee in ws.merged_cells.ranges:
      style_range(ws, str(rangee), border=border_thin)
  
  for cl in ['E5','K5','E31','E56','K56','E87','K87']:
    ws[cl].border = Border(top=thick,right=thick,bottom=thin)
  for cl in ['E6','K6','E32','E57','K57','E88','K88']:
    ws[cl].border = Border(top=thin,right=thick)
  for cl in ['C5','I5','C31','C56','I56','C87','I87']:
    ws[cl].border = Border(top=thick,right=thin)
  for cl in ['B6','H6','B32','B57','H57','B88','H88']:
    ws[cl].border = Border(left=thick)
  for cl in ['D22','J22','D48']:
    ws[cl].border = Border(right=thin,bottom=thick)
  
  for cl in ['B','H']:
    for n in [59,60,90,91]:
      plus=0
      while plus<=18 :
        ws[cl+str(n+plus)].border = Border(left=thick,right=thin)
        plus+=3
    
  for cl in ['E','K']:
    for n in [59,60,90,91]:
      plus=0
      while plus<=18:
        ws[cl+str(n+plus)].border = Border(left=thin,right=thick)
        plus+=3

  for cl in ['B78','H78','B109','H109']:
    ws[cl].border = Border(left=thick,bottom=thick,right=thin)
  for cl in ['E78','K78','E109','K109']:
    ws[cl].border = Border(right=thick,left=thin,bottom=thick)
  for cl in ['C78','I78','C109','I109']:
    ws[cl].border = Border(left=thin,bottom=thick)
  for cl in ['D78','J78','D109','J109']:
    ws[cl].border = Border(bottom=thick,right=thin)
  
  ####print area####
  ws.print_area = 'B5:K107'
  return "Yeahhhhhh"

def laborlist_v5(labor_df,dt,wb):
  alphabets = string.ascii_uppercase
  #page1
  ws = wb['근무일지_page1']
  t = ['월','화','수','목','금','토','일']
  r = datetime(int(dt[:4]),int(dt[5:7]),int(dt[8:10])).weekday()
  dt_format = dt[:4]+'. '+dt[5:7]+'. '+dt[8:10]+' '+t[r]+'요일'
  ws['B2'].value = '제 802 중대 근무일지\n'+'( '+dt_format+' )'
  ws['B2'].font = Font(name='굴림체',size=18,bold=True) 
  team_name_dict={}
  team_name_dict['1소대장']='경위 이일소'; team_name_dict['1소대부관']='경사 안일부'
  team_name_dict['2소대장']='경위 김이소'; team_name_dict['2소대부관']='경위 송이부'
  team_name_dict['3소대장']='경위 강삼소'; team_name_dict['3소대부관']='경사 오삼부'

  labor_subset = labor_df[lambda x : x['날짜']==dt].iloc[:,5:]
  labor_subset.reset_index(drop=True,inplace=True)
  temp_ls = list(labor_subset.values[0])
  today_labor = sorted(labor_subset.loc[:,[True if (e=='일') or (e=='당') else False for e in temp_ls]].columns)
  post_edit_ls = []
  for i,e in enumerate(today_labor) :
    if (len(today_labor) == 3) or (len(today_labor)==4) :
      if i != 0 :
        if today_labor[i-1][0] == today_labor[i][0]:
          ws['B'+str(14+i)].value = ''
          ws['D'+str(14+i)].value = team_name_dict[e]
          post_edit_ls.append('B'+str(14+i))
          post_edit_ls.append('C'+str(14+i))
        else :
          ws['B'+str(14+i)].value = e[0]+' '+e[1]+' '+e[2]
          ws['D'+str(14+i)].value = team_name_dict[e]
      else :
        ws['B'+str(14+i)].value = e[0]+' '+e[1]+' '+e[2]
        ws['D'+str(14+i)].value = team_name_dict[e]
    else :
      print(today_labor)
      raise ValueError('Labor is overloaded!!! Check this out!!!')

  thin = Side(border_style="thin",color="000000")
  thick= Side(border_style="thick",color="000000")
  border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)
  if len(today_labor)==4:
    ws['G17'].border = border_thin
    """ws.merge_cells('N11:N17')
    ws.merge_cells('O11:O17')"""

  for rangee in ws.merged_cells.ranges:
    tmp = str(rangee).split(':')[0]
    if tmp not in ['B2','B8','B19','B33','B39']:
      if len(today_labor)==4:
        style_range(ws, str(rangee), border=border_thin)
      else :
        if tmp[1:] != '17':
          style_range(ws, str(rangee), border=border_thin)
  
  allbutup_brdr = Border(left=Side(style="thin", color="000000"),
                  right=Side(style='thin', color="000000"),
                  bottom=Side(style="thin", color="000000"))
  onlyside_brdr = Border(left=Side(style="thin",color="000000"),
                         right=Side(style="thin",color="000000"))
  if len(today_labor)==4 :
    ws['N17'].border = allbutup_brdr
    ws['O17'].border = allbutup_brdr
    ws['N16'].border = onlyside_brdr
    ws['O16'].border = onlyside_brdr

  for cl in post_edit_ls:
    if cl[0] == 'B':
      ws[cl].border = Border(left=thin,bottom=thin)
      ws[cl[0]+str(int(cl[1:])-1)].border = Border(left=thin,top=thin)
    elif cl[0] == 'C':
      ws[cl].border = Border(bottom=thin,right=thin)
      ws[cl[0]+str(int(cl[1:])-1)].border = Border(right=thin,top=thin)

  # page2
  ws2 = wb['근무일지_page2']
  for rangee in ws2.merged_cells.ranges:
    tmp = str(rangee).split(':')[0]
    if tmp not in ['B2','B15','B22','B36']:
      style_range(ws2, str(rangee), border=border_thin)

  # page3
  ws3 = wb['근무일지_page3']
  for rangee in ws3.merged_cells.ranges:
    tmp = str(rangee).split(':')[0]
    if tmp not in ['B2','B27','B28','B36']:
      style_range(ws3, str(rangee), border=border_thin)

  # page4,5,6
  ws4 = wb['근무일지_page4']
  ws5 = wb['근무일지_page5']
  ws6 = wb['근무일지_page6']
  for rangee in ws4.merged_cells.ranges:
    tmp = str(rangee).split(':')[0]
    if tmp not in ['B2','B5']: 
      style_range(ws4, str(rangee), border=border_thin)
  for rangee in ws5.merged_cells.ranges:
    tmp = str(rangee).split(':')[0]
    if tmp not in ['B2']: 
      style_range(ws5, str(rangee), border=border_thin)
  for rangee in ws6.merged_cells.ranges:
    tmp = str(rangee).split(':')[0]
    if tmp not in ['B2']: 
      style_range(ws6, str(rangee), border=border_thin)

  # page4 editing
  
  for i in range(6,44):
    ws4['P'+str(i)].border = Border(left=thick)
    ws4['A'+str(i)].border = Border(right=thick)
  for i in range(3,43):
    ws5['P'+str(i)].border = Border(left=thick)
    ws6['P'+str(i)].border = Border(left=thick)
    ws5['A'+str(i)].border = Border(right=thick)
    ws6['A'+str(i)].border = Border(right=thick)
  for i in range(len(alphabets[1:15])):
    ws4[alphabets[i+1]+str(5)].border = Border(bottom=thick)
    ws5[alphabets[i+1]+str(2)].border = Border(bottom=thick)
    ws6[alphabets[i+1]+str(2)].border = Border(bottom=thick)
    ws4[alphabets[i+1]+str(44)].border = Border(top=thick)
    ws5[alphabets[i+1]+str(43)].border = Border(top=thick)
    ws6[alphabets[i+1]+str(43)].border = Border(top=thick)

  return 'LaborList is Done.'

def text_parsing(sub_df):
  cnt = len(sub_df)
  txt = ''
  cnt_ls=[0,0,0,0]
  if cnt == 0 :
    return txt, cnt, cnt_ls
  else :
    temp = sub_df[lambda x : x['소 대']=='1소대']
    if len(temp) != 0 :
      for nm in temp['성 명'].values:
        txt += ' '+nm
      txt += '(1P)'
      cnt_ls[0]=len(temp['성 명'].values)
    temp = sub_df[lambda x : x['소 대']=='2소대']
    if len(temp) != 0 :
      for nm in temp['성 명'].values:
        txt += ' '+nm
      txt += '(2P)'
      cnt_ls[1]=len(temp['성 명'].values)
    temp = sub_df[lambda x : x['소 대']=='3소대']
    if len(temp) != 0 :
      for nm in temp['성 명'].values:
        txt += ' '+nm
      txt += '(3P)'
      cnt_ls[2]=len(temp['성 명'].values)
    temp = sub_df[lambda x : x['소 대']=='본부소대']
    if len(temp) != 0 :
      for nm in temp['성 명'].values:
        txt += ' '+nm
      txt += '(HQ)'
      cnt_ls[3]=len(temp['성 명'].values)
    return txt, cnt, cnt_ls

def spacebar_adjust(strr):
  # 두 자릿수는 그대로 한 자릿수는 띄어쓰기 하나 해서 return
  if len(strr)==1 :
    return ' '+strr
  elif len(strr)==2 :
    return strr
  else :
    raise ValueError("Over 100 is not feasible..")

def worklist_p3_txt(sagoja_df,team):
  tmp_df = sagoja_df[lambda x : x['소 대']==team]
  if len(tmp_df)==0:
    return '•'
  elif len(tmp_df)==1:
    tmp = tmp_df['종 별'].values[0]
    if len(tmp)==2 :
      return tmp_df['성 명'].values[0] + '('+tmp[0]+'     '+tmp[1]+')'
    else:
      return tmp_df['성 명'].values[0] + '('+tmp+')'
  else :
    rst = ''
    for i in range(len(tmp_df)):
      if i != len(tmp_df)-1 :
        tmp = tmp_df['종 별'].values[i]
        if len(tmp)==2 :
          rst += tmp_df['성 명'].values[i] + '('+tmp[0]+'     '+tmp[1]+')'+'\n'
        else:
          rst += tmp_df['성 명'].values[i] + '('+tmp+')'+'\n'
      else :
        tmp = tmp_df['종 별'].values[i]
        if len(tmp)==2 :
          rst += tmp_df['성 명'].values[i] + '('+tmp[0]+'     '+tmp[1]+')'
        else:
          rst += tmp_df['성 명'].values[i] + '('+tmp+')'
  return rst

def validation_v1(sagoja_df,nm):
  if nm in sagoja_df['성 명'].values :
    raise ValueError("영외활동자인 대원입니다. 확인해주세요.")
  else :
    return

def worklist_p3_txt_janrew(ls,team):
  if len(ls) == 0:
    return '•'
  else :
    team_ls = []
    for e in (ls):
      if e.split('-')[0] == team:
        team_ls.append(e)
  if len(team_ls)==0:
    return '•'
  rst = ''
  for i,e in enumerate(team_ls) :
    tmp = e.split('-')
    if i != len(team_ls)-1:
      if len(tmp[1]) ==2 :
        rst += tmp[2]+'('+tmp[1][0]+'     '+tmp[1][1]+')'+'\n'
      else :
        rst += tmp[2]+'('+tmp[1]+')'+'\n'
    else :
      if len(tmp[1]) ==2 :
        rst += tmp[2]+'('+tmp[1][0]+'     '+tmp[1][1]+')'
      else :
        rst += tmp[2]+'('+tmp[1]+')'
  return rst

def worklist_p3_txt_v2(sagoja_df,key,team,IsStrip=True):
  tmp_df = sagoja_df[lambda x : x['소 대']==team].sort_values(by=key,ascending=True)
  #print(tmp_df)
  if len(tmp_df)==0:
    return '•'
  elif len(tmp_df)==1:
    tmp = tmp_df[key].values[0]
    if len(tmp)==2 and IsStrip:
      return tmp_df['성 명'].values[0] + '('+tmp[0]+'     '+tmp[1]+')'
    else:
      return tmp_df['성 명'].values[0] + '('+tmp+')'
  else :
    rst = ''
    for i in range(len(tmp_df)):
      if i != len(tmp_df)-1 :
        tmp = tmp_df[key].values[i]
        if len(tmp)==2 and IsStrip :
          rst += tmp_df['성 명'].values[i] + '('+tmp[0]+'     '+tmp[1]+')'+'\n'
        else:
          rst += tmp_df['성 명'].values[i] + '('+tmp+')'+'\n'
      else :
        tmp = tmp_df[key].values[i]
        if len(tmp)==2 and IsStrip :
          rst += tmp_df['성 명'].values[i] + '('+tmp[0]+'     '+tmp[1]+')'
        else:
          rst += tmp_df['성 명'].values[i] + '('+tmp+')'
  return rst

def worklist_p3_txt_v2_hq(sagoja_df,key,team,IsStrip=True):
  tmp_df = sagoja_df[lambda x : x['소대']==team].sort_values(by=key,ascending=False)
  if len(tmp_df)==0:
    return '•'
  elif len(tmp_df)==1:
    tmp = tmp_df[key].values[0]
    if len(tmp)==2 and IsStrip:
      return tmp_df['성명'].values[0] + '('+tmp[0]+'     '+tmp[1]+')'
    else:
      return tmp_df['성명'].values[0] + '('+tmp+')'
  else :
    rst = ''
    for i in range(len(tmp_df)):
      if i != len(tmp_df)-1 :
        tmp = tmp_df[key].values[i]
        if len(tmp)==2 and IsStrip :
          rst += tmp_df['성명'].values[i] + '('+tmp[0]+'     '+tmp[1]+')'+'\n'
        else:
          rst += tmp_df['성명'].values[i] + '('+tmp+')'+'\n'
      else :
        tmp = tmp_df[key].values[i]
        if len(tmp)==2 and IsStrip :
          rst += tmp_df['성명'].values[i] + '('+tmp[0]+'     '+tmp[1]+')'
        else:
          rst += tmp_df['성명'].values[i] + '('+tmp+')'
  return rst

def worklist_p3_txt_janrew_v2(ls,team,IsStrip=False):
  if len(ls) == 0:
    return '•'
  else :
    team_ls = []
    for e in (ls):
      if e.split('-')[0] == team:
        team_ls.append(e)
  if len(team_ls)==0:
    return '•'
  rst = ''
  for i,e in enumerate(team_ls) :
    tmp = e.split('-')
    if i != len(team_ls)-1:
      if len(tmp[1]) ==2 and IsStrip:
        rst += tmp[2]+'('+tmp[1][0]+'     '+tmp[1][1]+')'+'\n'
      else :
        rst += tmp[2]+'('+tmp[1]+')'+'\n'
    else :
      if len(tmp[1]) ==2 and IsStrip:
        rst += tmp[2]+'('+tmp[1][0]+'     '+tmp[1][1]+')'
      else :
        rst += tmp[2]+'('+tmp[1]+')'
  return rst

def spacebar_adjust_janrew(strr):
  if len(strr)==2:
    return strr[0]+'      '+strr[1]
  elif len(strr)==3:
    return strr[0]+'  '+strr[1]+'  '+strr[2]
  else :
    return strr

def length_extract(strr):
  if strr == '•':
    return 0
  else :
    return len(strr.split('\n'))

def validation_v2(sagoja_df,nm,IsNone=False):
  if IsNone:
    pass
  else :
    if nm=='':
      raise ValueError("대원을 입력하세요.")
  if nm in sagoja_df['성 명'].values :
    raise ValueError("영외활동자인 대원입니다. 확인해주세요.")
  else :
    return

def validation_team_v2(info,sinsang,nmonly=False):
  if info == '':
    return
  if nmonly :
    # info = 'nm-1'이런식
    if str(sinsang[lambda x : x['성명']==info.split('-')[0]]['소대'].values[0]) == info.split('-')[1]:
      return
    else :
      raise ValueError("대원과 소대가 일치하지 않습니다.")
  else :
    # info = '1소대-운전-오정훈'
    tmp = str(sinsang[lambda x : x['성명']==info.split('-')[-1]]['소대'].values[0])
    if info.split('-')[0][:2] == '본부':
      if tmp == 'HQ': return
      else : raise ValueError("대원과 소대가 일치하지 않습니다.")
    else :
      if tmp == info.split('-')[0][0]:return
      else : raise ValueError("대원과 소대가 일치하지 않습니다.")

def worklist_v11(sagoja_df,labor_df,sinsang, wb, dt):
  alphabets = string.ascii_uppercase
  thin = Side(border_style="thin",color="000000")
  thick = Side(border_style="thick", color="000000")
  double = Side(border_style="double",color="000000")

  ws = wb['업무보고_page1']
  sinsang = sinsang[['성명','보직','기수','소대','전역일자','전입일자']]
  temp_dt = dt[2:4]+'.'+dt[5:7]+'.'+dt[8:10]
  sinsang = sinsang[lambda x : x['전역일자']>temp_dt]
  sinsang = sinsang[lambda x : x['전입일자']<=temp_dt]
  ws['F10'] = '   '+str(9)+'\n'+'           '+spacebar_adjust(str(len(sinsang)))
  ws['I10'] = '   '+str(2)+'\n'+'           '+spacebar_adjust(str(len(sinsang[lambda x:x['소대']==1])))
  ws['L10'] = '   '+str(2)+'\n'+'           '+spacebar_adjust(str(len(sinsang[lambda x:x['소대']==2])))
  ws['O10'] = '   '+str(2)+'\n'+'           '+spacebar_adjust(str(len(sinsang[lambda x:x['소대']==3])))
  ws['R10'] = '   '+str(3)+'\n'+'           '+spacebar_adjust(str(len(sinsang[lambda x:x['소대']=='HQ'])))

  t = ['월','화','수','목','금','토','일']
  r = datetime(int(dt[:4]),int(dt[5:7]),int(dt[8:10])).weekday()
  input_dt = dt[:4]+'년 '+dt[5:7]+'. '+dt[8:10]+' '+t[r]+'요일'
  ws['B5'] = '【'+input_dt+'】'+'                                      【802의무경찰대】'
  heuga = sagoja_df[lambda x :x['종 별']=='정기휴가']
  whybak = sagoja_df[lambda x :x['종 별']=='정기외박']
  specialbak = sagoja_df[lambda x :x['종 별']=='특별외박']
  sickbak = sagoja_df[lambda x :x['종 별']=='병가']
  edu = sagoja_df[lambda x :x['종 별']=='교육']
  chungwon = sagoja_df[lambda x :x['종 별']=='청원휴가']
  guitar = sagoja_df[(sagoja_df['종 별']=='휴직') | (sagoja_df['종 별']=='공가')]#guitar = sagoja_df[lambda x :x['종 별']=='기타']
  heuga_info = text_parsing(heuga)
  whybak_info = text_parsing(whybak)
  specialbak_info = text_parsing(specialbak)
  sickbak_info = text_parsing(sickbak)
  edu_info = text_parsing(edu)
  chungwon_info = text_parsing(chungwon)
  guitar_info = text_parsing(guitar)
  total_ls = []
  ws['D22'] = heuga_info[0]; ws['U22'] = heuga_info[1]; total_ls.append(heuga_info[2])
  ws['D23'] = whybak_info[0]; ws['U23'] = whybak_info[1]; total_ls.append(whybak_info[2])
  ws['D24'] = specialbak_info[0]; ws['U24'] = specialbak_info[1]; total_ls.append(specialbak_info[2])
  ws['D25'] = sickbak_info[0]; ws['U25'] = sickbak_info[1]; total_ls.append(sickbak_info[2])
  ws['D26'] = edu_info[0]; ws['U26'] = edu_info[1]; total_ls.append(edu_info[2])
  ws['D27'] = chungwon_info[0]; ws['U27'] = chungwon_info[1]; total_ls.append(chungwon_info[2])
  ws['D28'] = guitar_info[0]; ws['U28'] = guitar_info[1]; total_ls.append(guitar_info[2])
  ws['U29'] = len(sagoja_df)

  for ii,nb in enumerate(range(13,20)):
    for ij,ap in enumerate(['I','L','O','R']):
      if total_ls[ii][ij] != 0:
        ws[ap+str(nb)] = total_ls[ii][ij]
      else :
        ws[ap+str(nb)] = ''
    ws['F'+str(nb)] = sum(total_ls[ii])
  
  # 지휘요원 처리
  leader_dict = labor_df[lambda x : x['날짜']==dt].iloc[0,2:].to_dict()
  # 중대장님,행소님,반장님은 휴가가 아닌 이상 일근처리 -> 그냥 '휴무'면 '일근'처리하자.
  if leader_dict['행정소대장'] == '휴':
    leader_dict['행정소대장'] = '일'
  if leader_dict['행정부관'] == '휴':
    leader_dict['행정부관'] = '일'
  if leader_dict['중대장'] == '휴':
    leader_dict['중대장'] = '일'
  dang=ill=heu=bee=vac=edu=sick=0
  dang_str=ill_str=heu_str=bee_str=vac_str=edu_str=sick_str=''
  for ix, lp in enumerate(leader_dict.keys()):
    if leader_dict[lp] == '당':
      dang +=1
      if dang_str == '':
        dang_str += lp.replace('소대부관','부소대장').replace('부관','부소대장')
      else :
        dang_str += '\n'+lp.replace('소대부관','부소대장').replace('부관','부소대장')
    elif leader_dict[lp] == '일':
      ill += 1
      if ill_str == '':
        ill_str += lp.replace('소대부관','부소대장').replace('부관','부소대장')
      else :
        ill_str += '\n'+lp.replace('소대부관','부소대장').replace('부관','부소대장')
    elif leader_dict[lp] == '휴':
      heu += 1
      if heu_str == '':
        heu_str += lp.replace('소대부관','부소대장').replace('부관','부소대장')
      else :
        heu_str += '\n'+lp.replace('소대부관','부소대장').replace('부관','부소대장')
    elif leader_dict[lp] == '비':
      bee += 1
      if bee_str == '':
        bee_str += lp.replace('소대부관','부소대장').replace('부관','부소대장')
      else :
        bee_str += '\n'+lp.replace('소대부관','부소대장').replace('부관','부소대장')
    elif leader_dict[lp] == '연':
      vac += 1
      if vac_str == '':
        vac_str += lp.replace('소대부관','부소대장').replace('부관','부소대장')
      else :
        vac_str += '\n'+lp.replace('소대부관','부소대장').replace('부관','부소대장')
    elif leader_dict[lp] == '교':
      edu += 1
      if edu_str == '':
        edu_str += lp.replace('소대부관','부소대장').replace('부관','부소대장')
      else :
        edu_str += '\n'+lp.replace('소대부관','부소대장').replace('부관','부소대장')
    elif leader_dict[lp] == '병':
      sick += 1
      if sick_str == '':
        sick_str += lp.replace('소대부관','부소대장').replace('부관','부소대장')
      else :
        sick_str += '\n'+lp.replace('소대부관','부소대장').replace('부관','부소대장')
    else :
      raise ValueError("Check this out.")
  ws['B33'] = dang_str; ws['E33'] = ill_str; ws['H33'] = heu_str; ws['K33'] = bee_str
  ws['N33'] = vac_str; ws['Q33'] = edu_str; ws['T33'] = sick_str
  
  nb1=nb2=nb3=nbHQ=0
  #병가나 교육도 사실 사고자로 치긴 해야 할텐데
  for i in heu_str.split('\n')+bee_str.split('\n')+vac_str.split('\n')+edu_str.split('\n')+sick_str.split('\n'):
    try:
      if i[0] == '1' : nb1 += 1
      elif i[0] == '2' : nb2 += 1
      elif i[0] == '3' : nb3 += 1
      elif i[0] == '행' or i[0] == '중' : nbHQ += 1
      else : raise ValueError("Check this out..")
    except :
      pass
  ws['I11'] = '   '+str(nb1)+'\n'+'           '+spacebar_adjust(str(len(sagoja_df[lambda x : x['소 대']=='1소대'])))
  ws['L11'] = '   '+str(nb2)+'\n'+'           '+spacebar_adjust(str(len(sagoja_df[lambda x : x['소 대']=='2소대'])))
  ws['O11'] = '   '+str(nb3)+'\n'+'           '+spacebar_adjust(str(len(sagoja_df[lambda x : x['소 대']=='3소대'])))
  ws['R11'] = '   '+str(nbHQ)+'\n'+'           '+spacebar_adjust(str(len(sagoja_df[lambda x : x['소 대']=='본부소대'])))
  ws['F11'] = '   '+str(nb1+nb2+nb3+nbHQ)+'\n'+'           '+spacebar_adjust(str(len(sagoja_df)))
  ws['I12'] = '   '+str(2-nb1)+'\n'+'           '+spacebar_adjust(str(int(ws['I10'].value.split(' ')[-1])-len(sagoja_df[lambda x : x['소 대']=='1소대'])))
  ws['L12'] = '   '+str(2-nb2)+'\n'+'           '+spacebar_adjust(str(int(ws['L10'].value.split(' ')[-1])-len(sagoja_df[lambda x : x['소 대']=='2소대'])))
  ws['O12'] = '   '+str(2-nb3)+'\n'+'           '+spacebar_adjust(str(int(ws['O10'].value.split(' ')[-1])-len(sagoja_df[lambda x : x['소 대']=='3소대'])))
  ws['R12'] = '   '+str(3-nbHQ)+'\n'+'           '+spacebar_adjust(str(int(ws['R10'].value.split(' ')[-1])-len(sagoja_df[lambda x : x['소 대']=='본부소대'])))
  ws['F12'] = '   '+str(9-nb1-nb2-nb3-nbHQ)+'\n'+'           '+spacebar_adjust(str(int(ws['F10'].value.split(' ')[-1])-len(sagoja_df)))

  # page1 border design
  save_merged_border(ws)
  worklist_p1_design_v2(ws)
  
  #################################
  ws2 = wb['업무보고_page2']
  save_merged_border(ws2)
  
  guard_1p = input("1소대 입초근무자 : ")
  validation_v2(sagoja_df,guard_1p,False)
  validation_team_v2(str(guard_1p+'-'+'1'),sinsang,True)
  guard_2p = input("2소대 입초근무자 : ")
  validation_v2(sagoja_df,guard_2p,False)
  validation_team_v2(str(guard_2p+'-'+'2'),sinsang,True)
  guard_3p = input("3소대 입초근무자 : ")
  validation_v2(sagoja_df,guard_3p,False)
  validation_team_v2(str(guard_3p+'-'+'3'),sinsang,True)
  stay_member = []
  print("잔류자 입력을 시작합니다. 같은 사유의 잔류자가 여럿일 경우 3소대-잔류-백상아 마승환 <-으로 띄어쓰기를 해주세요.")
  while True:
    tmp = input("잔류자 : 소대-사유-성명 를 입력하세요. EX)2소대-운전-김민재. 없으면 Enter를 누르세요. ")
    nm_tmp = tmp.split('-')[-1].split(' ')
    if len(nm_tmp) == 1 :
      validation_v2(sagoja_df,tmp.split('-')[-1],True)
    else :
      for p in nm_tmp :
        validation_v2(sagoja_df,p,True)
    if tmp == '':
      break
    else :
      stay_member.append(tmp)
  worklist_p2_design_v1(ws2,len(stay_member))###########
  
  ws2['J22'] = guard_1p; ws2['J23'] = guard_2p; ws2['J24'] = guard_3p
  hangjung_txt =''; chisa_txt=''
  hangjung = sinsang[(sinsang['소대']=='HQ') & (sinsang['보직']=='행정')]['성명'].values
  chisa = sinsang[(sinsang['소대']=='HQ') & (sinsang['보직']=='취사')]['성명'].values
  for hj in hangjung:
    if hj not in sagoja_df['성 명'].values:
      hangjung_txt += ' '+hj
  for cs in chisa:
    if cs not in sagoja_df['성 명'].values:
      chisa_txt += ' '+cs
  ws2['N30'] = hangjung_txt; ws2['N31'] = chisa_txt
  ws2['AE30'] = str(len(hangjung_txt.split(' '))-1); ws2['AE31'] = str(len(chisa_txt.split(' '))-1);
  guitar_nb1=guitar_nb2=guitar_nb3=0
  if len(stay_member) == 0 :
    ws2['B32'] = '총     계'; ws2['AE32'] = str(int(ws2['AE30'].value)+int(ws2['AE31'].value))
  else :
    for i in range(len(stay_member)):
      tmp = stay_member[i].split('-')
      ws2['B'+str(32+i)] = tmp[0][0]+'  '+tmp[0][1]+'  '+tmp[0][2]
      ws2['H'+str(32+i)] = spacebar_adjust_janrew(tmp[1])
      ws2['N'+str(32+i)] = ' '+tmp[2]
      ws2['AE'+str(32+i)] = len(tmp[-1].split(' ')) #1 #이거 바꿔야지..merge cell parsing후에
      if tmp[0][0]=='1': guitar_nb1+=len(tmp[-1].split(' ')) #1
      elif tmp[0][0]=='2': guitar_nb2+=len(tmp[-1].split(' ')) #1
      elif tmp[0][0]=='3': guitar_nb3+=len(tmp[-1].split(' ')) #1
      else : raise ValueError("Check this out...")
    ws2['B'+str(32+len(stay_member))] = '총     계'
    cnt = 0
    for i in range(30,32+len(stay_member)):
      cnt += int(ws2['AE'+str(i)].value)
    ws2['AE'+str(32+len(stay_member))] = str(cnt)
  ws2['AA4']=str(guitar_nb1); ws2['AA5']=str(guitar_nb2); ws2['AA6']=str(guitar_nb3)

  ################################
  
  ws2['G4'] = len(sinsang[lambda x : x['소대']==1])
  ws2['G5'] = len(sinsang[lambda x : x['소대']==2])
  ws2['G6'] = len(sinsang[lambda x : x['소대']==3])
  ws2['G10'] = len(sinsang[lambda x : x['소대']=='HQ'])
  ws2['K4'] = str(ws['I11'].value.split(' ')[-1])
  ws2['K5'] = str(ws['L11'].value.split(' ')[-1])
  ws2['K6'] = str(ws['O11'].value.split(' ')[-1])
  ws2['L10'] = str(ws['R11'].value.split(' ')[-1])
  ws2['O4'] = str(int(ws2['G4'].value) - int(ws2['K4'].value))
  ws2['O5'] = str(int(ws2['G5'].value) - int(ws2['K5'].value))
  ws2['O6'] = str(int(ws2['G6'].value) - int(ws2['K6'].value))
  #복귀예정자
  comback_1 = comback_2 = comback_3 = comback_hq = 0
  another_dt = '~ '+dt[5:7]+'.'+dt[8:10]
  lastday_df = sagoja_df[sagoja_df['기 간'].str.contains(another_dt)]
  comback_1 = len(lastday_df[lambda x : x['소 대']=='1소대'])
  comback_2 = len(lastday_df[lambda x : x['소 대']=='2소대'])
  comback_3 = len(lastday_df[lambda x : x['소 대']=='3소대'])
  comback_hq = len(lastday_df[lambda x : x['소 대']=='본부소대'])
  ws2['S4'] = comback_1; ws2['S5'] = comback_2; ws2['S6'] = comback_3; ws2['Q10'] = comback_hq
  ws2['AE4'] = str(int(ws2['O4'].value) - 1 - int(ws2['AA4'].value))
  ws2['AE5'] = str(int(ws2['O5'].value) - 1 - int(ws2['AA5'].value))
  ws2['AE6'] = str(int(ws2['O6'].value) - 1 - int(ws2['AA6'].value))
  chuldong=2; bigo = '운전(1), 무전(1)'
  hq_chuldong = sinsang[(sinsang['소대']=='HQ') & (sinsang['보직']=='운전') | (sinsang['보직']=='무전')]['성명'].values
  hq_hangchi = sinsang[(sinsang['소대']=='HQ') & (sinsang['보직']=='행정') | (sinsang['보직']=='취사')]['성명'].values                                               
  for hq_agent in hq_chuldong:
    if hq_agent in sagoja_df[lambda x : x['소 대']=='본부소대']['성 명'].values:
      chuldong -= 1
      if sinsang[lambda x : x['성명']==hq_agent]['보직'].values[0] == '무전':
        bigo = '운전(1)'
      else :
        bigo = '무전(1)'
  if chuldong == 0 :
    bigo = ''
    print("Why both hq_agents are out..? out of mind..?")
  ws2['V10'] = str(chuldong); ws2['AA10'] = bigo;
  ws2['I13'] = ' '+str(1)+'\n'+'        '+spacebar_adjust(ws2['AE4'].value)
  ws2['N13'] = ' '+str(1)+'\n'+'        '+spacebar_adjust(ws2['AE5'].value)
  ws2['S13'] = ' '+str(1)+'\n'+'        '+spacebar_adjust(ws2['AE6'].value)
  ws2['X13'] = ' '+str(1 if leader_dict['중대장']!='연' else 0)+'\n'+'        '+spacebar_adjust(ws2['V10'].value)
  chuldong_daewon = str(int(ws2['AE4'].value)+int(ws2['AE5'].value)+int(ws2['AE6'].value)+chuldong)
  ws2['AC13'] = '  '+str(3+ (1 if leader_dict['중대장']!='연' else 0))+'\n'+'          '+spacebar_adjust(chuldong_daewon)

  ############################3
  ws3 = wb['업무보고_page3']
  save_merged_border(ws3)
  worklist_p3_design_v1(ws3)
  g5_v = [s.strip(' ') for s in ws2['I13'].value.split('\n')]
  ws3['G5'] = g5_v[0]+'/'+g5_v[1]
  m5_v = [s.strip(' ') for s in ws2['N13'].value.split('\n')]
  ws3['M5'] = m5_v[0]+'/'+m5_v[1]
  s5_v = [s.strip(' ') for s in ws2['S13'].value.split('\n')]
  ws3['S5'] = s5_v[0]+'/'+s5_v[1]
  y5_v = [s.strip(' ') for s in ws2['X13'].value.split('\n')]
  ws3['Y5'] = y5_v[0]+'/'+y5_v[1]
  ae5_v = [s.strip(' ') for s in ws2['AC13'].value.split('\n')]
  ws3['AE5'] = ae5_v[0]+'/'+ae5_v[1]

  ws3['G8'] = worklist_p3_txt_v2(sagoja_df,'종 별','1소대',False)
  ws3['M8'] = worklist_p3_txt_v2(sagoja_df,'종 별','2소대',False)
  ws3['S8'] = worklist_p3_txt_v2(sagoja_df,'종 별','3소대',False)
  ws3['Y8'] = worklist_p3_txt_v2(sagoja_df,'종 별','본부소대',False)
  ws3['AE8'] = str(len(sagoja_df))
  ws3['G25'] = guard_1p; ws3['M25'] = guard_2p; ws3['S25'] = guard_3p
  ws3['G31'] = worklist_p3_txt_janrew_v3(stay_member,'1소대',True)
  ws3['M31'] = worklist_p3_txt_janrew_v3(stay_member,'2소대',True)
  ws3['S31'] = worklist_p3_txt_janrew_v3(stay_member,'3소대',True)

  whychul_member = []
  whychul_nmonly = []
  print("외출자 입력을 시작합니다. 해당 대원들에 대해 한 명씩 입력하세요.")
  while True:
    tmp = input("외출자 : 소대-사유-성명 를 입력하세요. EX)본부소대-특별-인대영. 없으면 Enter를 누르세요. ")
    validation_v2(sagoja_df,tmp.split('-')[-1],True)
    validation_team_v2(tmp,sinsang,False)
    if tmp == '':
      break
    else :
      whychul_member.append(tmp)
      whychul_nmonly.append(tmp.split('-')[-1])
  ws3['G21'] = worklist_p3_txt_janrew_v2(whychul_member,'1소대',False)
  ws3['M21'] = worklist_p3_txt_janrew_v2(whychul_member,'2소대',False)
  ws3['S21'] = worklist_p3_txt_janrew_v2(whychul_member,'3소대',False)
  ws3['Y21'] = worklist_p3_txt_janrew_v2(whychul_member,'본부소대',False)
  ws3['AE21'] = str(len(whychul_member))
  hq_janrew = sinsang[(sinsang['소대']=='HQ') & (sinsang['보직']=='행정') | (sinsang['보직']=='취사')]
  hq_janrew = hq_janrew[~hq_janrew['성명'].isin(sagoja_df['성 명'].values)]
  hq_janrew = hq_janrew[~hq_janrew['성명'].isin(whychul_nmonly)]
  ws3['Y31']=worklist_p3_txt_v2_hq(hq_janrew,'보직','HQ',False)
  ws3['AE31'] = str(length_extract(ws3['G31'].value)+length_extract(ws3['M31'].value)+length_extract(ws3['S31'].value)+len(hq_janrew))

  ws3['G41'] = str(length_extract(ws3['G8'].value)+length_extract(ws3['G21'].value)+1+length_extract(ws3['G31'].value))+'/'+ws['I10'].value.split(' ')[-1]
  ws3['M41'] = str(length_extract(ws3['M8'].value)+length_extract(ws3['M21'].value)+1+length_extract(ws3['M31'].value))+'/'+ws['L10'].value.split(' ')[-1]
  ws3['S41'] = str(length_extract(ws3['S8'].value)+length_extract(ws3['S21'].value)+1+length_extract(ws3['S31'].value))+'/'+ws['O10'].value.split(' ')[-1]
  ws3['Y41'] = str(length_extract(ws3['Y8'].value)+length_extract(ws3['Y21'].value)+0+length_extract(ws3['Y31'].value))+'/'+ws['R10'].value.split(' ')[-1]
  ws3['AE41'] = str(int(ws3['AE8'].value)+int(ws3['AE21'].value)+int(3)+int(0)+int(ws3['AE31'].value))+'/'+ws['F10'].value.split(' ')[-1]
  return "Done"

def save_merged_border(ws):
  for merged_cells in ws.merged_cells.ranges:
    style = ws.cell(merged_cells.min_row, merged_cells.min_col)._style
    for col in range(merged_cells.min_col, merged_cells.max_col + 1):
        for row in range(merged_cells.min_row, merged_cells.max_row + 1): 
            ws.cell(row, col)._style = style
  return

def worklist_p1_design_v2(ws):
  # page1 border design
  alphabets = string.ascii_uppercase
  thin = Side(border_style="thin",color="000000")
  thick = Side(border_style="thick", color="000000")
  double = Side(border_style="double",color="000000")
  medium = Side(border_style="medium",color="000000")
  all_brdr = Border(left=Side(style="thin", color="000000"),
                  right=Side(style='thin', color="000000"),
                  top=Side(style="thin", color="000000"),
                  bottom=Side(style="thin", color="000000"),
                  diagonal=Side(style="thin", color="000000"),
                  diagonal_direction=0,
                  outline=Side(style="thin", color="000000"),
                  vertical=Side(style="thin", color="000000"),
                  horizontal=Side(style="thin", color="000000"))

  r38 = [alphabets[1:-4][i]+'38' for i in range(len(alphabets[1:-4]))]
  for cl in r38 :
    ws[cl].border = Border(top=medium)
  cW = ['W6','W7'] + ['W'+str(i) for i in range(9,20)] + ['W'+str(i) for i in range(22,30)] + ['W'+str(i) for i in range(32,38)]
  for cl in cW :
    ws[cl].border = Border(left=medium)  
  ws['E6'].border = Border(right=thin,bottom=double)
  ws['E7'].border = Border(right=thin,top=double)
  ws['E9'].border = Border(right=double,bottom=thin,top=medium)
  ws['B19'].border = Border(left=medium,bottom=medium)
  ws['C19'].border = Border(bottom=medium,right=double)
  for cl in ['C'+str(i) for i in range(14,19)]:
    ws[cl].border = Border(right=double)
  ws['C13'].border = Border(right=double,top=thin)
  ws['C29'].border = Border(right=double,bottom=medium,top=thin)
  ws['A6'].border = Border(right=medium)
  ws['A7'].border = Border(right=medium)
  ws['A9'].border = Border(right=medium)
  for cl in ['A'+str(i) for i in range(13,20)]:
    ws[cl].border = Border(right=medium)
  ws['B20'].border = Border(top=medium); ws['C20'].border = Border(top=medium)
  for cl in ['B8','C8','D8','E8']:
    ws[cl].border = Border(top=medium)
  ws['A29'].border = Border(right=medium)
  return

def worklist_p2_design_v1(ws,nm_janrew):
  # page2 border design
  alphabets = string.ascii_uppercase
  thin = Side(border_style="thin",color="000000")
  thick = Side(border_style="thick", color="000000")
  double = Side(border_style="double",color="000000")
  medium = Side(border_style="medium",color="000000")
  all_brdr = Border(left=Side(style="thin", color="000000"),
                  right=Side(style='thin', color="000000"),
                  top=Side(style="thin", color="000000"),
                  bottom=Side(style="thin", color="000000"),
                  diagonal=Side(style="thin", color="000000"),
                  diagonal_direction=0,
                  outline=Side(style="thin", color="000000"),
                  vertical=Side(style="thin", color="000000"),
                  horizontal=Side(style="thin", color="000000"))

  for cl in ['AI'+str(i) for i in [3,4,5,6]+[9,10]]:
    ws[cl].border = Border(left=medium)
  for cl in ['AI'+str(i) for i in [12,13]+[22,23,24,25]+[i for i in range(29,33+nm_janrew)]]:
    ws[cl].border = Border(left=thin)
  ws['AE3'].border = Border(left=thin,top=medium,bottom=double)
  ws['V9'].border = Border(left=thin,top=medium,bottom=double)
  ws['V10'].border = Border(left=thin,top=double,bottom=medium)
  ws['AE29'].border = Border(left=thin,top=thin,bottom=thin)
  for rg in ws.merged_cells.ranges:
    for rw in [str(int(32)+i) for i in range(nm_janrew+1)]:
      if rw in str(rg).split(':')[0]:
        style_range(ws,str(rg),border=all_brdr)
  return

def worklist_p3_design_v1(ws):
  # page3 border design
  alphabets = string.ascii_uppercase
  thin = Side(border_style="thin",color="000000")
  thick = Side(border_style="thick", color="000000")
  double = Side(border_style="double",color="000000")
  medium = Side(border_style="medium",color="000000")
  all_brdr = Border(left=Side(style="thin", color="000000"),
                  right=Side(style='thin', color="000000"),
                  top=Side(style="thin", color="000000"),
                  bottom=Side(style="thin", color="000000"),
                  diagonal=Side(style="thin", color="000000"),
                  diagonal_direction=0,
                  outline=Side(style="thin", color="000000"),
                  vertical=Side(style="thin", color="000000"),
                  horizontal=Side(style="thin", color="000000"))
  dotted = Side(border_style="dotted",color="000000")

  for cl in ['AI'+str(i) for i in range(4,44)]:
    ws[cl].border = Border(left=thin)
  r44=[]
  for i in range(len(alphabets[1:])):
    r44.append(alphabets[1:][i])
  for i in ['A'+ap for ap in alphabets[:8]]:
    r44.append(i)
  for cl in r44:
    ws[cl+str(44)].border = Border(top=thin)    
  
  for cl in ['A'+str(i) for i in range(41,44)]:
    ws[cl].border = Border(right=thin)
  for cl in ['F','L','R','X','AD']:
    for n in [42,43]:
      ws[cl+str(n)].border = Border(right=thin)
  for cl in r44:
    if cl in ['F','L','R','X','AD']:
      ws[cl+str(41)].border = Border(top=dotted,right=thin)
    else :
      ws[cl+str(41)].border = Border(top=dotted)
  return

def worklist_p3_txt_janrew_v3(ls,team,IsStrip=False):
  if len(ls) == 0:
    return '•'
  else :
    team_ls = []
    for e in (ls):
      if e.split('-')[0] == team:
        team_ls.append(e)
  if len(team_ls)==0:
    return '•'
  rst = ''
  reset_ls = []
  for e in team_ls :
    tmp = e.split('-')
    if len(tmp[2].split(' '))==1:
      reset_ls.append(e)
    else :
      nm_tmp = tmp[2].split(' ')
      for i in range(len(nm_tmp)):
        reset_ls.append(tmp[0]+'-'+tmp[1]+'-'+nm_tmp[i])

  for i,e in enumerate(reset_ls) :
    tmp = e.split('-')
    if i != len(reset_ls)-1:
      if len(tmp[1]) ==2 and IsStrip:
        rst += tmp[2]+'('+tmp[1][0]+'     '+tmp[1][1]+')'+'\n'
      else :
        rst += tmp[2]+'('+tmp[1]+')'+'\n'
    else :
      if len(tmp[1]) ==2 and IsStrip:
        rst += tmp[2]+'('+tmp[1][0]+'     '+tmp[1][1]+')'
      else :
        rst += tmp[2]+'('+tmp[1]+')'
  return rst

def PatrolChart_v6(dt,wb):
  t = ['월','화','수','목','금','토','일']
  r = datetime(int(dt[:4]),int(dt[5:7]),int(dt[8:10])).weekday()
  input_dt = dt[:4]+'년 '+dt[5:7]+'월 '+dt[8:10]+'일'+'\n'+t[r]+'요일'
  ws = wb['순찰표']
  col_dims = ws.column_dimensions
  row_dims = ws.row_dimensions

  ws['B5'].value = '( 지 하 ) 순찰표\n'+input_dt
  ws['H5'].value = '( 청 사 앞 ) 순찰표\n'+input_dt
  ws['B25'].value = '( 서 측 계 단 ) 순찰표\n'+input_dt
  ws['B45'].value = '(3소대 내무실) 순찰표\n'+input_dt
  ws['H45'].value = '(2소대 내무실) 순찰표\n'+input_dt
  ws['B65'].value = '(1소대 내무실) 순찰표\n'+input_dt
  ws['H65'].value = '(본부소대 내무실) 순찰표\n'+input_dt

  thick = Side(border_style="thick", color="000000")
  thin = Side(border_style="thin",color="000000")
  border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)
  border_thick = Border(top=thick, left=thick, right=thick, bottom=thick)
  double = Side(border_style="double", color="000000")

  for rangee in ws.merged_cells.ranges:
      style_range(ws, str(rangee), border=border_thin)

  # row border
  for ap in ['A','G']:
    if ap == 'A':
      for n in range(5,83):
        if n not in [23,24,43,44,63,64]:
          ws[ap+str(n)].border = Border(right=thick)
    else :
      for n in list(range(5,23))+list(range(45,63))+list(range(65,83)):
        ws[ap+str(n)].border = Border(right=thick)
  for ap in ['F','L']:
    if ap == 'F':
      for n in range(5,83):
        if n not in [23,24,43,44,63,64]:
          ws[ap+str(n)].border = Border(left=thick)
    else :
      for n in list(range(5,23))+list(range(45,63))+list(range(65,83)):
        ws[ap+str(n)].border = Border(left=thick)
  # col border
  for n in [4,24,44,64]:
    if n != 24 :
      for ap in ['B','C','D','E','H','I','J','K']:
        ws[ap+str(n)].border = Border(bottom=thick)
    else :
      for ap in ['B','C','D','E']:
        ws[ap+str(n)].border = Border(bottom=thick)
  for n in [23,43,63,83]:
    if n != 43 :
      for ap in ['B','C','D','E','H','I','J','K']:
        ws[ap+str(n)].border = Border(top=thick)
    else :
      for ap in ['B','C','D','E']:
        ws[ap+str(n)].border = Border(top=thick)
  
  ####print area####
  ws.print_area ='A3:L83'
  return "Yeahhhhhh"

def Sagoja_v6(table, dt, db,IsCompress=True):
    sub_table = table[(table['StartDate']<=dt) & (table['EndDate']>=dt)]
    rst_df = pd.DataFrame(columns=['연번','구분','소대','소대','성명','기간','종별','비고'])
    sub_table.reset_index(inplace=True,drop=True)
    sub_table['Team'] = sub_table['Team'].map(lambda x : x + '소대' if len(x)==1 else '본부소대')

    type_ls = []
    for i in range(len(sub_table)):
    #구분
        tpcls = sub_table.loc[i,'Class']
        if '휴가' in tpcls :
            type_ls.append('휴 가')
        elif '정기외박' in tpcls :
            type_ls.append('외 박')
        elif '특별외박' in tpcls :
            type_ls.append('특 박')
        elif '병가' in tpcls :
            type_ls.append('병 가')
        elif '교육' in tpcls :
            type_ls.append('교 육')
        elif '휴직' in tpcls :
            type_ls.append('휴 직')
        elif '공가' in tpcls :
            type_ls.append('공 가')
        else :
            type_ls.append('기 타')
            print(" Problem : ", tpcls)
            raise ValueError('Check this out !!!')
    sub_table['구 분'] = type_ls
    sub_table.rename(columns={'Name':'성 명','Team':'소 대','Class':'종 별','Specialty':'비 고','기간':'기 간'},inplace=True)
    sub_table.drop(['StartDate','EndDate'],axis=1,inplace=True)
    sub_table = sub_table[['구 분','소 대','성 명','기 간','종 별','비 고']]
    if IsCompress == True :
        pass
    else :
        full_sagoja = sub_table.copy()
    rst_table = pd.DataFrame(columns = ['구 분','소 대','성 명','기 간','종 별','비 고'])
    heuga_df = SortByTeam(sub_table[lambda x : x['구 분']=='휴 가'])
    try:
        heuga_df = Compress_v1(heuga_df)
        heuga_df.iloc[1:,0] = ''
    except:
        pass
  
    whybak_df = SortByTeam(sub_table[lambda x : x['구 분']=='외 박'])
    try:
        whybak_df = Compress_v1(whybak_df)
        whybak_df.iloc[1:,0] = ''
    except:
        pass
    specialbak_df = SortByTeam(sub_table[lambda x : x['구 분']=='특 박'])
    try:
        specialbak_df = Compress_v1(specialbak_df)
        specialbak_df.iloc[1:,0] = ''
    except:
        pass
    sickbak_df = SortByTeam(sub_table[lambda x : x['구 분']=='병 가'])
    try:
        sickbak_df = Compress_v1(sickbak_df)
        sickbak_df.iloc[1:,0] = ''
    except:
        pass
    edubak_df = SortByTeam(sub_table[lambda x : x['구 분']=='교 육'])
    try :
        edubak_df = Compress_v1(edubak_df)
        edubak_df.iloc[1:,0] = ''
    except :
        pass
    heuzic_df = SortByTeam(sub_table[lambda x : x['구 분']=='휴 직'])
    try :
        heuzic_df = Compress_v1(heuzic_df)
        heuzic_df.iloc[1:,0] = ''
    except :
        pass
    gongga_df = SortByTeam(sub_table[lambda x : x['구 분']=='공 가'])
    try :
        gongga_df = Compress_v1(gongga_df)
        gongga_df.iloc[1:,0] = ''
    except :
        pass
    guitar_df = SortByTeam(sub_table[lambda x : x['구 분']=='기 타'])
    try:
        guitar_df = Compress_v1(guitar_df)
        guitar_df.iloc[1:,0] = ''
    except:
        pass
    rst_table = pd.concat([rst_table,heuga_df])
    rst_table = pd.concat([rst_table,whybak_df])
    rst_table = pd.concat([rst_table,specialbak_df])
    rst_table = pd.concat([rst_table,sickbak_df])
    rst_table = pd.concat([rst_table,edubak_df])
    rst_table = pd.concat([rst_table,heuzic_df])
    rst_table = pd.concat([rst_table,gongga_df])
    rst_table = pd.concat([rst_table,guitar_df])
    rst_table.reset_index(drop=True,inplace=True)
    yeonbeon_ls = [i+1 for i in rst_table.index]
    rst_table['연 번'] = yeonbeon_ls
    if IsCompress == True :
        return rst_table[['연 번','구 분','소 대','성 명','기 간','종 별','비 고']]
    else :
        full_sagoja['연 번'] = yeonbeon_ls
        return full_sagoja[['연 번','구 분','소 대','성 명','기 간','종 별','비 고']]

def Simulator_v7(dt):
    filename_dt = dt[:4]+dt[5:7]+dt[8:10]
    df = pd.read_excel(glob.glob('*외박파트*')[0])
    df = df.loc[:,:'사고자']
    df.drop(['사고자'],axis=1,inplace=True)
    df_index = [str(idx)[:10] for idx in df.날짜.values]
    df.index = df_index
    df.drop(['요일','날짜'],axis=1, inplace=True)
    db = pd.read_excel(glob.glob('*신상표*')[0])
    book = openpyxl.load_workbook(glob.glob('*외박파트*')[0])
    total = openpyxl.load_workbook(glob.glob('*template*')[0])
    raw = book.active
    total_sheet = total.active
    ref = date_column_cell(df)
    alphabets = string.ascii_uppercase
    labor_df = pd.read_excel(glob.glob('*지휘요원*')[0])

    # step1 : extracting allpart
    allpart = Stacking_v3(df,db,raw,ref)
    # step2 : extracting sagoja table   
    sagoja_df = Sagoja_v6(allpart,dt,db,True)
    sagoja = SagojaDesign_v3(sagoja_df,dt,wb=total)
    # step3 : making checklist using sagoja table
    sagoja_df_F = Sagoja_v6(allpart,dt,db,False)
    sagoja_df_F_fltrd = sagoja_df_F[(sagoja_df_F['종 별'] != '휴직') & (sagoja_df_F['종 별'] != '공가')]
    checklist = checklist_v5(sagoja_df_F_fltrd,dt,wb=total)
    # step4 : making patrol chart given dt
    patrolchart = PatrolChart_v6(dt,wb=total)
    # step5 : making labor list given dt
    laborlist = laborlist_v5(labor_df,dt,wb=total)
    # step6 : making work report
    workreport = worklist_v11(sagoja_df_F,labor_df,db,total,dt)
    #fname = 'drive/My Drive/full_documents_%s.xlsx'%(filename_dt)
    total.save('full_documents_%s.xlsx'%(filename_dt))
    return "FINISHED!!!"

if __name__ == "__main__":
    print("====================================================================================================")
    print("802 Automation version 2 is starting now..")
    print("Please insert a date as a string form like 'YYYY-MM-DD'.")
    argument = sys.argv
    del argument[0]
    print('Date you requested : {}'.format(argument[0]))
    Simulator_v7(argument[0])
    print('Successfully finished !')
    print("====================================================================================================")